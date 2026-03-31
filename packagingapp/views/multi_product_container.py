from __future__ import annotations

import io
from typing import Dict, List

from django.conf import settings
from django.http import JsonResponse, HttpRequest, HttpResponse
from django.shortcuts import render, get_object_or_404
from django.views.decorators.http import require_GET, require_POST
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from packagingapp.models import PackagingCatalogue, ProductCatalogue, PackagingMaterial, Product
from packagingapp.utils.box_selection.engine import compute_max_quantity_only, run_mode1_and_render


def _product_dims(p: Product):
    return (float(p.product_length), float(p.product_width), float(p.product_height))


def _product_rotations(p: Product):
    r1 = 1 if getattr(p, "rotation_1", True) else 0
    r2 = 1 if getattr(p, "rotation_2", False) else 0
    r3 = 1 if getattr(p, "rotation_3", False) else 0
    if r1 == 0 and r2 == 0 and r3 == 0:
        r1 = 1
    return r1, r2, r3


def _product_label(p: Product) -> str:
    if p.product_name:
        return str(p.product_name)
    if p.product_id:
        return str(p.product_id)
    return str(p)


def _serialize_product_catalogues() -> List[Dict]:
    catalogues = ProductCatalogue.objects.all().order_by("name")
    payload = []

    for c in catalogues:
        products = c.products.all().order_by("-created_at")
        payload.append(
            {
                "id": c.id,
                "name": c.name,
                "description": c.description or "",
                "picture_url": c.picture.url if c.picture else "",
                "product_count": products.count(),
                "rows": [
                    {
                        "id": p.id,
                        "product_id": p.product_id or "",
                        "product_name": p.product_name or "",
                        "product_length": float(p.product_length),
                        "product_width": float(p.product_width),
                        "product_height": float(p.product_height),
                        "rotation_1": bool(p.rotation_1),
                        "rotation_2": bool(p.rotation_2),
                        "rotation_3": bool(p.rotation_3),
                        "weight": float(p.weight) if p.weight is not None else None,
                        "desired_qty": int(getattr(p, "desired_qty", 1) or 1),
                        "product_volume": float(p.product_volume) if p.product_volume is not None else None,
                        "picture_url": p.product_picture.url if p.product_picture else "",
                    }
                    for p in products
                ],
            }
        )
    return payload


def _serialize_packaging_catalogues() -> List[Dict]:
    catalogues = PackagingCatalogue.objects.all().order_by("name")
    payload = []

    for c in catalogues:
        materials = c.materials.all().order_by("part_number")
        payload.append(
            {
                "id": c.id,
                "name": c.name,
                "description": c.description or "",
                "picture_url": c.picture.url if c.picture else "",
                "material_count": materials.count(),
                "rows": [
                    {
                        "id": m.id,
                        "part_number": m.part_number,
                        "part_description": m.part_description,
                        "packaging_type": m.packaging_type,
                        "branding": m.branding,
                        "part_length": float(m.part_length),
                        "part_width": float(m.part_width),
                        "part_height": float(m.part_height),
                        "external_length": float(m.external_length) if m.external_length is not None else None,
                        "external_width": float(m.external_width) if m.external_width is not None else None,
                        "external_height": float(m.external_height) if m.external_height is not None else None,
                        "part_weight": float(m.part_weight) if m.part_weight is not None else None,
                        "part_volume": float(m.part_volume) if m.part_volume is not None else None,
                    }
                    for m in materials
                ],
            }
        )
    return payload


def _rank_top5_for_product(product: Product, materials: List[PackagingMaterial]) -> List[Dict]:
    desired_qty = max(int(getattr(product, "desired_qty", 1) or 1), 1)

    product_dims = _product_dims(product)
    r1, r2, r3 = _product_rotations(product)

    product_vol = product_dims[0] * product_dims[1] * product_dims[2]
    scored: List[Dict] = []

    for m in materials:
        container = (float(m.part_length), float(m.part_width), float(m.part_height))
        max_qty = compute_max_quantity_only(product_dims, container, r1, r2, r3)

        if max_qty >= desired_qty:
            container_vol = float(m.part_volume) if m.part_volume is not None else (container[0] * container[1] * container[2])
            usage = (desired_qty * product_vol) / container_vol if container_vol > 0 else 0.0

            scored.append(
                {
                    "material_obj": m,
                    "material": {
                        "id": m.id,
                        "part_number": m.part_number,
                        "part_description": m.part_description,
                        "packaging_type": m.packaging_type,
                        "branding": m.branding,
                        "part_length": float(m.part_length),
                        "part_width": float(m.part_width),
                        "part_height": float(m.part_height),
                        "external_length": float(m.external_length) if m.external_length is not None else None,
                        "external_width": float(m.external_width) if m.external_width is not None else None,
                        "external_height": float(m.external_height) if m.external_height is not None else None,
                        "part_weight": float(m.part_weight) if m.part_weight is not None else None,
                        "part_volume": float(m.part_volume) if m.part_volume is not None else None,
                    },
                    "max_qty": int(max_qty),
                    "usage": float(usage),
                    "container_vol": float(container_vol),
                }
            )

    scored.sort(key=lambda x: (-x["usage"], x["container_vol"]))
    return scored[:5]


@require_GET
def multi_product_container_selection(request: HttpRequest) -> HttpResponse:
    context = {
        "product_catalogues": ProductCatalogue.objects.all().order_by("name"),
        "packaging_catalogues": PackagingCatalogue.objects.all().order_by("name"),
        "product_catalogues_payload": _serialize_product_catalogues(),
        "packaging_catalogues_payload": _serialize_packaging_catalogues(),
    }
    return render(request, "multi_product_container/multi_product_container_selection.html", context)


@require_POST
def multi_product_container_run(request: HttpRequest) -> JsonResponse:
    product_catalogue_id = request.POST.get("product_catalogue_id")
    packaging_catalogue_id = request.POST.get("packaging_catalogue_id")

    if not product_catalogue_id or not packaging_catalogue_id:
        return JsonResponse({"ok": False, "error": "Missing catalogue selection."}, status=400)

    product_catalogue = get_object_or_404(ProductCatalogue, id=product_catalogue_id)
    packaging_catalogue = get_object_or_404(PackagingCatalogue, id=packaging_catalogue_id)

    products = list(product_catalogue.products.all().order_by("-created_at"))
    materials = list(PackagingMaterial.objects.filter(catalogue_id=packaging_catalogue.id).order_by("part_number"))

    results: List[Dict] = []

    for p in products:
        top5 = _rank_top5_for_product(p, materials)
        results.append(
            {
                "product": {
                    "id": p.id,
                    "label": _product_label(p),
                    "product_id": p.product_id or "",
                    "product_name": p.product_name or "",
                    "product_length": float(p.product_length),
                    "product_width": float(p.product_width),
                    "product_height": float(p.product_height),
                    "rotation_1": bool(p.rotation_1),
                    "rotation_2": bool(p.rotation_2),
                    "rotation_3": bool(p.rotation_3),
                    "desired_qty": int(getattr(p, "desired_qty", 1) or 1),
                    "weight": float(p.weight) if p.weight is not None else None,
                    "product_volume": float(p.product_volume) if p.product_volume is not None else None,
                    "picture_url": p.product_picture.url if p.product_picture else "",
                },
                "top5": [
                    {
                        "material": row["material"],
                        "max_qty": row["max_qty"],
                        "usage": row["usage"],
                    }
                    for row in top5
                ],
            }
        )

    return JsonResponse({"ok": True, "results": results})


@require_POST
def multi_product_container_draw(request: HttpRequest) -> JsonResponse:
    product_id = request.POST.get("product_id")
    container_id = request.POST.get("container_id")

    if not product_id or not container_id:
        return JsonResponse({"ok": False, "error": "Missing product_id or container_id."}, status=400)

    p = get_object_or_404(Product, id=product_id)
    m = get_object_or_404(PackagingMaterial, id=container_id)

    desired_qty = max(int(getattr(p, "desired_qty", 1) or 1), 1)

    product_dims = _product_dims(p)
    container_dims = (float(m.part_length), float(m.part_width), float(m.part_height))
    r1, r2, r3 = _product_rotations(p)

    result = run_mode1_and_render(
        product_dims,
        container_dims,
        r1, r2, r3,
        settings.MEDIA_ROOT,
        draw_limit=desired_qty,
    )

    return JsonResponse(
        {
            "ok": True,
            "image_url": settings.MEDIA_URL + result.image_rel_path,
            "max_quantity": int(result.max_quantity),
            "desired_qty": desired_qty,
        }
    )


@require_POST
def multi_product_container_export_excel(request: HttpRequest) -> HttpResponse:
    product_catalogue_id = request.POST.get("product_catalogue_id")
    packaging_catalogue_id = request.POST.get("packaging_catalogue_id")

    if not product_catalogue_id or not packaging_catalogue_id:
        return HttpResponse("Missing catalogue selection.", status=400)

    product_catalogue = get_object_or_404(ProductCatalogue, id=product_catalogue_id)
    packaging_catalogue = get_object_or_404(PackagingCatalogue, id=packaging_catalogue_id)

    products = list(product_catalogue.products.all().order_by("-created_at"))
    materials = list(PackagingMaterial.objects.filter(catalogue_id=packaging_catalogue.id).order_by("part_number"))

    wb = Workbook()
    ws_best = wb.active
    ws_best.title = "Best container"
    ws_full = wb.create_sheet("Full analysis")

    ws_best.append([
        "product_id", "product_name",
        "product_length", "product_width", "product_height",
        "desired_qty",
        "best_part_number", "best_part_description",
        "container_length", "container_width", "container_height",
        "usage", "max_qty",
        "packaging_type", "branding",
    ])

    ws_full.append([
        "product_id", "product_name",
        "product_length", "product_width", "product_height",
        "desired_qty",
        "rank",
        "part_number", "part_description",
        "container_length", "container_width", "container_height",
        "usage", "max_qty",
        "packaging_type", "branding",
    ])

    for p in products:
        top5 = _rank_top5_for_product(p, materials)

        if top5:
            best = top5[0]
            m = best["material"]

            ws_best.append([
                p.product_id, p.product_name,
                float(p.product_length), float(p.product_width), float(p.product_height),
                int(getattr(p, "desired_qty", 1) or 1),
                m["part_number"], m["part_description"],
                float(m["part_length"]), float(m["part_width"]), float(m["part_height"]),
                float(best["usage"]), int(best["max_qty"]),
                m["packaging_type"], m["branding"],
            ])
        else:
            ws_best.append([
                p.product_id, p.product_name,
                float(p.product_length), float(p.product_width), float(p.product_height),
                int(getattr(p, "desired_qty", 1) or 1),
                "", "", "", "", "", "", "", "", "",
            ])

        for idx, row in enumerate(top5, start=1):
            m = row["material"]
            ws_full.append([
                p.product_id, p.product_name,
                float(p.product_length), float(p.product_width), float(p.product_height),
                int(getattr(p, "desired_qty", 1) or 1),
                idx,
                m["part_number"], m["part_description"],
                float(m["part_length"]), float(m["part_width"]), float(m["part_height"]),
                float(row["usage"]), int(row["max_qty"]),
                m["packaging_type"], m["branding"],
            ])

    for ws in (ws_best, ws_full):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"multi_product_container_{product_catalogue.name}_{packaging_catalogue.name}.xlsx".replace(" ", "_")

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response