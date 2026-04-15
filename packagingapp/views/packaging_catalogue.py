from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
from openpyxl import Workbook

from packagingapp.access import (
    LOGIN_REQUIRED_MESSAGE,
    can_manage_packaging_catalogue,
    get_manageable_packaging_catalogue_or_404,
    get_visible_packaging_catalogue_or_404,
    visible_packaging_catalogues,
)
from packagingapp.forms import (
    PackagingCatalogueForm,
    PackagingMaterialForm,
    ExcelUploadForm,
    DrawingUploadForm,
    PackagingMaterialFilterForm,
)
from packagingapp.services.excel_import import import_packaging_excel
from packagingapp.services.drawing_import import import_drawings_zip


def catalogue_list(request):
    catalogues = visible_packaging_catalogues(request.user).order_by("-created_at")
    return render(
        request,
        "packaging_catalogue/catalogue_list.html",
        {"catalogues": catalogues},
    )


@login_required
def create_catalogue(request):
    if request.method == "POST":
        form = PackagingCatalogueForm(request.POST, request.FILES)
        if form.is_valid():
            catalogue = form.save(commit=False)
            catalogue.owner = request.user
            catalogue.is_public = False
            catalogue.save()
            messages.success(request, "Private packaging catalogue created successfully.")
            return redirect("catalogue_detail", pk=catalogue.pk)
    else:
        form = PackagingCatalogueForm()

    return render(
        request,
        "packaging_catalogue/create_catalogue.html",
        {"form": form},
    )


def catalogue_detail(request, pk):
    catalogue = get_visible_packaging_catalogue_or_404(request.user, pk=pk)
    materials = catalogue.materials.all().order_by("part_number")

    form = PackagingMaterialFilterForm(request.GET or None)

    if form.is_valid():
        if form.cleaned_data.get("part_number"):
            materials = materials.filter(part_number__icontains=form.cleaned_data["part_number"])

        if form.cleaned_data.get("part_description"):
            materials = materials.filter(part_description__icontains=form.cleaned_data["part_description"])

        if form.cleaned_data.get("packaging_type"):
            materials = materials.filter(packaging_type__icontains=form.cleaned_data["packaging_type"])

        if form.cleaned_data.get("branding"):
            materials = materials.filter(branding__icontains=form.cleaned_data["branding"])

        if form.cleaned_data.get("min_length"):
            materials = materials.filter(part_length__gte=form.cleaned_data["min_length"])
        if form.cleaned_data.get("max_length"):
            materials = materials.filter(part_length__lte=form.cleaned_data["max_length"])

        if form.cleaned_data.get("min_width"):
            materials = materials.filter(part_width__gte=form.cleaned_data["min_width"])
        if form.cleaned_data.get("max_width"):
            materials = materials.filter(part_width__lte=form.cleaned_data["max_width"])

        if form.cleaned_data.get("min_height"):
            materials = materials.filter(part_height__gte=form.cleaned_data["min_height"])
        if form.cleaned_data.get("max_height"):
            materials = materials.filter(part_height__lte=form.cleaned_data["max_height"])

        if form.cleaned_data.get("min_weight"):
            materials = materials.filter(part_weight__gte=form.cleaned_data["min_weight"])
        if form.cleaned_data.get("max_weight"):
            materials = materials.filter(part_weight__lte=form.cleaned_data["max_weight"])

        if form.cleaned_data.get("min_volume"):
            materials = materials.filter(part_volume__gte=form.cleaned_data["min_volume"])
        if form.cleaned_data.get("max_volume"):
            materials = materials.filter(part_volume__lte=form.cleaned_data["max_volume"])

    paginator = Paginator(materials, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "packaging_catalogue/catalogue_detail.html",
        {
            "catalogue": catalogue,
            "form": form,
            "page_obj": page_obj,
            "can_manage_catalogue": can_manage_packaging_catalogue(request.user, catalogue),
        },
    )


@login_required
@require_POST
def delete_catalogue(request, pk):
    catalogue = get_manageable_packaging_catalogue_or_404(request.user, pk=pk)
    catalogue.delete()
    messages.success(request, "Packaging catalogue deleted successfully.")
    return redirect("catalogue_list")


@login_required
def add_material(request, pk):
    catalogue = get_manageable_packaging_catalogue_or_404(request.user, pk=pk)

    if request.method == "POST":
        form = PackagingMaterialForm(request.POST, request.FILES)
        if form.is_valid():
            material = form.save(commit=False)
            material.catalogue = catalogue
            material.save()
            messages.success(request, "Item added successfully.")
            return redirect("catalogue_detail", pk=catalogue.pk)
    else:
        form = PackagingMaterialForm()

    return render(
        request,
        "packaging_catalogue/add_material.html",
        {"catalogue": catalogue, "form": form},
    )


@login_required
def upload_excel(request, pk):
    catalogue = get_manageable_packaging_catalogue_or_404(request.user, pk=pk)

    if request.method == "POST":
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                created_count, updated_count = import_packaging_excel(request.FILES["file"], catalogue)
                return render(
                    request,
                    "packaging_catalogue/upload_excel.html",
                    {
                        "form": ExcelUploadForm(),
                        "catalogue": catalogue,
                        "success": f"Upload completed. Created: {created_count}. Updated: {updated_count}.",
                    },
                )
            except Exception as e:
                return render(
                    request,
                    "packaging_catalogue/upload_excel.html",
                    {"form": form, "catalogue": catalogue, "error": str(e)},
                )
    else:
        form = ExcelUploadForm()

    return render(
        request,
        "packaging_catalogue/upload_excel.html",
        {"form": form, "catalogue": catalogue},
    )


@login_required
def upload_drawings_for_catalogue(request, pk):
    catalogue = get_manageable_packaging_catalogue_or_404(request.user, pk=pk)

    if request.method == "POST":
        form = DrawingUploadForm(request.POST, request.FILES)
        if form.is_valid():
            imported, not_matched, skipped = import_drawings_zip(
                request.FILES["zip_file"],
                catalogue,
            )
            return render(
                request,
                "packaging_catalogue/upload_drawings.html",
                {
                    "form": DrawingUploadForm(),
                    "catalogue": catalogue,
                    "success": f"Done. Imported: {imported}. Not matched: {not_matched}. Skipped: {skipped}.",
                },
            )
    else:
        form = DrawingUploadForm()

    return render(
        request,
        "packaging_catalogue/upload_drawings.html",
        {"form": form, "catalogue": catalogue},
    )


def download_excel_template(request, pk):
    catalogue = get_visible_packaging_catalogue_or_404(request.user, pk=pk)

    wb = Workbook()

    ws = wb.active
    ws.title = "Template"
    ws.append([
        "part_number",
        "part_description",
        "packaging_type",
        "branding",
        "packaging_materials",
        "part_length",
        "part_width",
        "part_height",
        "external_length",
        "external_width",
        "external_height",
        "part_weight",
    ])
    ws.append([
        "100001",
        "Sample corrugated box",
        "BOX",
        "Brand1",
        "Paperboard",
        300,
        200,
        150,
        310,
        210,
        160,
        1.250,
    ])

    info = wb.create_sheet(title="Instructions")
    info["A1"] = "Packaging Catalogue Upload Template"
    info["A2"] = "Units"
    info["B2"] = "Metric only"
    info["A3"] = "Internal dimensions"
    info["B3"] = "part_length, part_width, part_height in mm"
    info["A4"] = "External dimensions"
    info["B4"] = "external_length, external_width, external_height in mm"
    info["A5"] = "Weight"
    info["B5"] = "part_weight in kg"
    info["A6"] = "Allowed packaging_type"
    info["B6"] = "BOX, PALLET, CRATE, BAG"
    info["A7"] = "Catalogue"
    info["B7"] = catalogue.name

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"packaging_catalogue_template_{catalogue.pk}.xlsx"

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def edit_catalogue(request, pk):
    catalogue = get_manageable_packaging_catalogue_or_404(request.user, pk=pk)

    if request.method == "POST":
        form = PackagingCatalogueForm(request.POST, request.FILES, instance=catalogue)
        if form.is_valid():
            updated = form.save(commit=False)
            updated.owner = catalogue.owner
            updated.is_public = catalogue.is_public
            updated.save()
            messages.success(request, "Packaging catalogue updated successfully.")
            return redirect("catalogue_detail", pk=catalogue.pk)
    else:
        form = PackagingCatalogueForm(instance=catalogue)

    return render(
        request,
        "packaging_catalogue/edit_catalogue.html",
        {
            "catalogue": catalogue,
            "form": form,
        },
    )


def export_catalogue_excel(request, pk):
    catalogue = get_visible_packaging_catalogue_or_404(request.user, pk=pk)
    materials = catalogue.materials.all().order_by("part_number")

    wb = Workbook()
    ws = wb.active
    ws.title = "Packaging Catalogue Export"

    ws.append([
        "part_number",
        "part_description",
        "packaging_type",
        "branding",
        "packaging_materials",
        "part_length",
        "part_width",
        "part_height",
        "external_length",
        "external_width",
        "external_height",
        "part_weight",
        "part_volume",
        "drawing",
    ])

    for m in materials:
        ws.append([
            m.part_number,
            m.part_description,
            m.packaging_type,
            m.branding,
            m.packaging_materials,
            float(m.part_length) if m.part_length is not None else None,
            float(m.part_width) if m.part_width is not None else None,
            float(m.part_height) if m.part_height is not None else None,
            float(m.external_length) if m.external_length is not None else None,
            float(m.external_width) if m.external_width is not None else None,
            float(m.external_height) if m.external_height is not None else None,
            float(m.part_weight) if m.part_weight is not None else None,
            float(m.part_volume) if m.part_volume is not None else None,
            m.drawing.url if m.drawing else "",
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"packaging_catalogue_export_{catalogue.pk}.xlsx"

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
