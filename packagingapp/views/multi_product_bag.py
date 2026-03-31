from __future__ import annotations

from typing import Any, Dict, List

import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from django.conf import settings
from django.http import JsonResponse, HttpRequest, HttpResponse
from django.shortcuts import render, get_object_or_404
from django.views.decorators.http import require_GET, require_POST

from packagingapp.models import PackagingCatalogue, ProductCatalogue, PackagingMaterial, Product
from packagingapp.utils.bag_selection.engine import (
    build_required_bag_options,
    best_usage_for_bag,
    run_bag_mode1_and_render,
)


def _product_dims(p: Product):
    return (float(p.product_length), float(p.product_width), float(p.product_height))


def _product_label(p: Product) -> str:
    if p.product_name:
        return str(p.product_name)
    if p.product_id:
        return str(p.product_id)
    return str(p)


def _bag_dims_from_material(m: PackagingMaterial):
    """
    Correct BAG rule for this project:
    - bag length = part_length
    - bag width  = part_width
    """
    return (float(m.part_length), float(m.part_width))


def _serialize_product_catalogues() -> List[Dict[str, Any]]:
    catalogues = ProductCatalogue.objects.all().order_by("name")
    payload: List[Dict[str, Any]] = []

    for c in catalogues:
        products = c.products.all().order_by("-created_at")
        payload.append(
            {
                "id": c.id,
                "name": c.name,
                "description": getattr(c, "description", "") or "",
                "picture_url": c.picture.url if getattr(c, "picture", None) else "",
                "product_count": products.count(),
                "rows": [
                    {
                        "id": p.id,
                        "product_id": p.product_id or "",
                        "product_name": p.product_name or "",
                        "product_length": float(p.product_length),
                        "product_width": float(p.product_width),
                        "product_height": float(p.product_height),
                        "rotation_1": bool(p.rotation_1),
                        "rotation_2": bool(p.rotation_2),
                        "rotation_3": bool(p.rotation_3),
                        "weight": float(p.weight) if p.weight is not None else None,
                        "desired_qty": int(getattr(p, "desired_qty", 1) or 1),
                        "product_volume": float(p.product_volume) if getattr(p, "product_volume", None) is not None else None,
                        "picture_url": p.product_picture.url if p.product_picture else "",
                    }
                    for p in products
                ],
            }
        )
    return payload


def _serialize_packaging_catalogues() -> List[Dict[str, Any]]:
    catalogues = PackagingCatalogue.objects.all().order_by("name")
    payload: List[Dict[str, Any]] = []

    for c in catalogues:
        materials = c.materials.filter(packaging_type="BAG").order_by("part_number")
        payload.append(
            {
                "id": c.id,
                "name": c.name,
                "description": c.description or "",
                "picture_url": c.picture.url if getattr(c, "picture", None) else "",
                "material_count": materials.count(),
                "rows": [
                    {
                        "id": m.id,
                        "part_number": m.part_number,
                        "part_description": m.part_description,
                        "packaging_type": m.packaging_type,
                        "branding": m.branding,
                        "part_length": float(m.part_length),
                        "part_width": float(m.part_width),
                        "part_height": float(m.part_height),
                        "external_length": float(m.external_length) if getattr(m, "external_length", None) is not None else None,
                        "external_width": float(m.external_width) if getattr(m, "external_width", None) is not None else None,
                        "external_height": float(m.external_height) if getattr(m, "external_height", None) is not None else None,
                        "part_weight": float(m.part_weight) if getattr(m, "part_weight", None) is not None else None,
                        "part_volume": float(m.part_volume) if m.part_volume is not None else None,
                    }
                    for m in materials
                ],
            }
        )
    return payload


def _rank_top5_for_product(product: Product, materials: List[PackagingMaterial]) -> List[Dict[str, Any]]:
    desired_qty = max(int(getattr(product, "desired_qty", 1) or 1), 1)
    product_dims = _product_dims(product)

    bag_build = build_required_bag_options(
        product_l=product_dims[0],
        product_w=product_dims[1],
        product_h=product_dims[2],
        desired_qty=desired_qty,
    )

    required_bags = bag_build["required"]
    smooth_qty = int(bag_build["smooth_qty"])

    scored: List[Dict[str, Any]] = []

    for m in materials:
        bag_len, bag_w = _bag_dims_from_material(m)
        best = best_usage_for_bag(bag_len, bag_w, required_bags)

        if best:
            bag_area = bag_len * bag_w
            scored.append(
                {
                    "material": {
                        "id": m.id,
                        "part_number": m.part_number,
                        "part_description": m.part_description,
                        "packaging_type": m.packaging_type,
                        "branding": m.branding,
                        "part_length": float(m.part_length),
                        "part_width": float(m.part_width),
                        "part_height": float(m.part_height),
                        "external_length": float(m.external_length) if getattr(m, "external_length", None) is not None else None,
                        "external_width": float(m.external_width) if getattr(m, "external_width", None) is not None else None,
                        "external_height": float(m.external_height) if getattr(m, "external_height", None) is not None else None,
                        "part_weight": float(m.part_weight) if getattr(m, "part_weight", None) is not None else None,
                        "part_volume": float(m.part_volume) if m.part_volume is not None else None,
                    },
                    "bag_len": float(bag_len),
                    "bag_w": float(bag_w),
                    "usage": float(best["usage"]),
                    "best_required": [float(best["req_len"]), float(best["req_w"])],
                    "bag_area": float(bag_area),
                    "smooth_qty": smooth_qty,
                    "required_bags": [[float(x), float(y)] for (x, y) in required_bags],
                }
            )

    scored.sort(key=lambda x: (-x["usage"], x["bag_area"]))
    return scored[:5]


@require_GET
def multi_product_bag_selection(request: HttpRequest) -> HttpResponse:
    context = {
        "product_catalogues": ProductCatalogue.objects.all().order_by("name"),
        "packaging_catalogues": PackagingCatalogue.objects.all().order_by("name"),
        "product_catalogues_payload": _serialize_product_catalogues(),
        "packaging_catalogues_payload": _serialize_packaging_catalogues(),
    }
    return render(request, "multi_product_bag/multi_product_bag_selection.html", context)


@require_POST
def multi_product_bag_run(request: HttpRequest) -> JsonResponse:
    product_catalogue_id = request.POST.get("product_catalogue_id")
    packaging_catalogue_id = request.POST.get("packaging_catalogue_id")

    if not product_catalogue_id or not packaging_catalogue_id:
        return JsonResponse({"ok": False, "error": "Missing catalogue selection."}, status=400)

    product_catalogue = get_object_or_404(ProductCatalogue, id=product_catalogue_id)
    packaging_catalogue = get_object_or_404(PackagingCatalogue, id=packaging_catalogue_id)

    products = list(product_catalogue.products.all().order_by("-created_at"))
    materials = list(
        PackagingMaterial.objects.filter(
            catalogue_id=packaging_catalogue.id,
            packaging_type="BAG",
        ).order_by("part_number")
    )

    results: List[Dict[str, Any]] = []

    for p in products:
        top5 = _rank_top5_for_product(p, materials)
        results.append(
            {
                "product": {
                    "id": p.id,
                    "label": _product_label(p),
                    "product_id": p.product_id or "",
                    "product_name": p.product_name or "",
                    "product_length": float(p.product_length),
                    "product_width": float(p.product_width),
                    "product_height": float(p.product_height),
                    "desired_qty": max(int(getattr(p, "desired_qty", 1) or 1), 1),
                    "weight": float(p.weight) if p.weight is not None else None,
                    "product_volume": float(p.product_volume) if getattr(p, "product_volume", None) is not None else None,
                    "picture_url": p.product_picture.url if p.product_picture else "",
                },
                "top5": [
                    {
                        "material": row["material"],
                        "bag_len": row["bag_len"],
                        "bag_w": row["bag_w"],
                        "usage": row["usage"],
                        "best_required": row["best_required"],
                        "smooth_qty": row["smooth_qty"],
                        "required_bags": row["required_bags"],
                    }
                    for row in top5
                ],
            }
        )

    return JsonResponse({"ok": True, "results": results})


@require_POST
def multi_product_bag_draw(request: HttpRequest) -> JsonResponse:
    product_id = request.POST.get("product_id")
    bag_id = request.POST.get("bag_id")

    if not product_id or not bag_id:
        return JsonResponse({"ok": False, "error": "Missing product_id or bag_id."}, status=400)

    p = get_object_or_404(Product, id=product_id)
    m = get_object_or_404(PackagingMaterial, id=bag_id)

    if m.packaging_type != "BAG":
        return JsonResponse({"ok": False, "error": "Selected material is not a BAG."}, status=400)

    desired_qty = max(int(getattr(p, "desired_qty", 1) or 1), 1)
    product_dims = _product_dims(p)
    bag_len, bag_w = _bag_dims_from_material(m)

    bag_build = build_required_bag_options(
        product_l=product_dims[0],
        product_w=product_dims[1],
        product_h=product_dims[2],
        desired_qty=desired_qty,
    )

    required_bags = bag_build["required"]
    solutions = bag_build["solutions"]
    smooth_qty = int(bag_build["smooth_qty"])

    best = best_usage_for_bag(bag_len, bag_w, required_bags)
    if not best:
        return JsonResponse(
            {"ok": False, "error": "Selected bag does not fit any required option for this product."},
            status=400,
        )

    render_result = run_bag_mode1_and_render(
        product=product_dims,
        selected_bag=(bag_len, bag_w),
        desired_qty=desired_qty,
        solutions=solutions,
        media_root=settings.MEDIA_ROOT,
        draw_limit=desired_qty,
    )

    return JsonResponse(
        {
            "ok": True,
            "image_url": settings.MEDIA_URL + render_result.image_rel_path,
            "desired_qty": desired_qty,
            "smooth_qty": smooth_qty,
            "bag_len": bag_len,
            "bag_w": bag_w,
            "best_required": [float(best["req_len"]), float(best["req_w"])],
            "usage": float(best["usage"]),
            "used_layout": list(render_result.used_layout),
            "inner_box": [float(x) for x in render_result.inner_box],
            "required_bag": [float(x) for x in render_result.required_bag],
            "required_bags": [[float(x), float(y)] for (x, y) in required_bags],
        }
    )


@require_POST
def multi_product_bag_export_excel(request: HttpRequest) -> HttpResponse:
    product_catalogue_id = request.POST.get("product_catalogue_id")
    packaging_catalogue_id = request.POST.get("packaging_catalogue_id")

    if not product_catalogue_id or not packaging_catalogue_id:
        return HttpResponse("Missing catalogue selection.", status=400)

    product_catalogue = get_object_or_404(ProductCatalogue, id=product_catalogue_id)
    packaging_catalogue = get_object_or_404(PackagingCatalogue, id=packaging_catalogue_id)

    products = list(product_catalogue.products.all().order_by("-created_at"))
    materials = list(
        PackagingMaterial.objects.filter(
            catalogue_id=packaging_catalogue.id,
            packaging_type="BAG",
        ).order_by("part_number")
    )

    wb = Workbook()
    ws_best = wb.active
    ws_best.title = "Best bag"
    ws_full = wb.create_sheet("Full analysis")

    ws_best.append([
        "product_id", "product_name",
        "product_length", "product_width", "product_height",
        "desired_qty",
        "best_part_number", "best_part_description",
        "bag_length", "bag_width",
        "usage",
        "best_required_length", "best_required_width",
        "packaging_type", "branding",
    ])

    ws_full.append([
        "product_id", "product_name",
        "product_length", "product_width", "product_height",
        "desired_qty",
        "rank",
        "part_number", "part_description",
        "bag_length", "bag_width",
        "usage",
        "best_required_length", "best_required_width",
        "packaging_type", "branding",
    ])

    for p in products:
        top5 = _rank_top5_for_product(p, materials)

        if top5:
            best = top5[0]
            m = best["material"]
            ws_best.append([
                p.product_id, p.product_name,
                float(p.product_length), float(p.product_width), float(p.product_height),
                max(int(getattr(p, "desired_qty", 1) or 1), 1),
                m["part_number"], m["part_description"],
                float(best["bag_len"]), float(best["bag_w"]),
                float(best["usage"]),
                float(best["best_required"][0]), float(best["best_required"][1]),
                m["packaging_type"], m["branding"],
            ])
        else:
            ws_best.append([
                p.product_id, p.product_name,
                float(p.product_length), float(p.product_width), float(p.product_height),
                max(int(getattr(p, "desired_qty", 1) or 1), 1),
                "", "", "", "", "", "", "", "", "",
            ])

        for idx, row in enumerate(top5, start=1):
            m = row["material"]
            ws_full.append([
                p.product_id, p.product_name,
                float(p.product_length), float(p.product_width), float(p.product_height),
                max(int(getattr(p, "desired_qty", 1) or 1), 1),
                idx,
                m["part_number"], m["part_description"],
                float(row["bag_len"]), float(row["bag_w"]),
                float(row["usage"]),
                float(row["best_required"][0]), float(row["best_required"][1]),
                m["packaging_type"], m["branding"],
            ])

    for ws in (ws_best, ws_full):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"multi_product_bag_{product_catalogue.name}_{packaging_catalogue.name}.xlsx".replace(" ", "_")

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response