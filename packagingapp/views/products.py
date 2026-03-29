from io import BytesIO
import zipfile

from django.contrib import messages
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import Workbook

from packagingapp.forms import (
    ProductCatalogueForm,
    ProductForm,
    ProductExcelUploadForm,
    ProductImagesZipUploadForm,
    ProductFilterForm,
)
from packagingapp.models import ProductCatalogue, Product
from packagingapp.services.product_excel_import import import_product_excel


def product_catalogues(request):
    catalogues = ProductCatalogue.objects.all().order_by("-created_at")
    return render(
        request,
        "product_catalogue/catalogues.html",
        {"catalogues": catalogues},
    )


def create_product_catalogue(request):
    if request.method == "POST":
        form = ProductCatalogueForm(request.POST, request.FILES)
        if form.is_valid():
            catalogue = form.save()
            messages.success(request, "Product catalogue created successfully.")
            return redirect("product_catalogue_detail", catalogue_id=catalogue.pk)
    else:
        form = ProductCatalogueForm()

    return render(
        request,
        "product_catalogue/create_catalogue.html",
        {"form": form},
    )


def edit_product_catalogue(request, catalogue_id):
    catalogue = get_object_or_404(ProductCatalogue, pk=catalogue_id)

    if request.method == "POST":
        form = ProductCatalogueForm(request.POST, request.FILES, instance=catalogue)
        if form.is_valid():
            form.save()
            messages.success(request, "Product catalogue updated successfully.")
            return redirect("product_catalogue_detail", catalogue_id=catalogue.pk)
    else:
        form = ProductCatalogueForm(instance=catalogue)

    return render(
        request,
        "product_catalogue/edit_catalogue.html",
        {
            "catalogue": catalogue,
            "form": form,
        },
    )


def delete_product_catalogue(request, catalogue_id):
    catalogue = get_object_or_404(ProductCatalogue, id=catalogue_id)
    if request.method == "POST":
        catalogue.delete()
        messages.success(request, "Product catalogue deleted successfully.")
        return redirect("product_catalogues")
    return redirect("product_catalogues")


def product_catalogue_detail(request, catalogue_id):
    catalogue = get_object_or_404(ProductCatalogue, id=catalogue_id)
    products = catalogue.products.all().order_by("product_id")

    form = ProductFilterForm(request.GET or None)

    if form.is_valid():
        cd = form.cleaned_data

        if cd.get("product_id"):
            products = products.filter(product_id__icontains=cd["product_id"])

        if cd.get("product_name"):
            products = products.filter(product_name__icontains=cd["product_name"])

        if cd.get("min_length"):
            products = products.filter(product_length__gte=cd["min_length"])
        if cd.get("max_length"):
            products = products.filter(product_length__lte=cd["max_length"])

        if cd.get("min_width"):
            products = products.filter(product_width__gte=cd["min_width"])
        if cd.get("max_width"):
            products = products.filter(product_width__lte=cd["max_width"])

        if cd.get("min_height"):
            products = products.filter(product_height__gte=cd["min_height"])
        if cd.get("max_height"):
            products = products.filter(product_height__lte=cd["max_height"])

        if cd.get("min_weight"):
            products = products.filter(weight__gte=cd["min_weight"])
        if cd.get("max_weight"):
            products = products.filter(weight__lte=cd["max_weight"])

        if cd.get("min_desired_qty"):
            products = products.filter(desired_qty__gte=cd["min_desired_qty"])
        if cd.get("max_desired_qty"):
            products = products.filter(desired_qty__lte=cd["max_desired_qty"])

        if cd.get("min_volume"):
            products = products.filter(product_volume__gte=cd["min_volume"])
        if cd.get("max_volume"):
            products = products.filter(product_volume__lte=cd["max_volume"])

    paginator = Paginator(products, 25)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "product_catalogue/catalogue_detail.html",
        {
            "catalogue": catalogue,
            "form": form,
            "page_obj": page_obj,
        },
    )


def add_product(request, catalogue_id):
    catalogue = get_object_or_404(ProductCatalogue, id=catalogue_id)

    if request.method == "POST":
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save(commit=False)
            product.catalogue = catalogue
            product.save()
            messages.success(request, "Product added successfully.")
            return redirect("product_catalogue_detail", catalogue_id=catalogue.id)
    else:
        form = ProductForm()

    return render(
        request,
        "product_catalogue/add_product.html",
        {"catalogue": catalogue, "form": form},
    )


def upload_products_excel(request, catalogue_id):
    catalogue = get_object_or_404(ProductCatalogue, id=catalogue_id)

    if request.method == "POST":
        form = ProductExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                created_count, updated_count = import_product_excel(request.FILES["file"], catalogue)
                return render(
                    request,
                    "product_catalogue/upload_excel.html",
                    {
                        "form": ProductExcelUploadForm(),
                        "catalogue": catalogue,
                        "success": f"Upload completed. Created: {created_count}. Updated: {updated_count}.",
                    },
                )
            except Exception as e:
                return render(
                    request,
                    "product_catalogue/upload_excel.html",
                    {"form": form, "catalogue": catalogue, "error": str(e)},
                )
    else:
        form = ProductExcelUploadForm()

    return render(
        request,
        "product_catalogue/upload_excel.html",
        {"catalogue": catalogue, "form": form},
    )


def download_product_excel_template(request, catalogue_id):
    catalogue = get_object_or_404(ProductCatalogue, pk=catalogue_id)

    wb = Workbook()

    ws = wb.active
    ws.title = "Template"
    ws.append([
        "product_id",
        "product_name",
        "product_length",
        "product_width",
        "product_height",
        "rotation_1",
        "rotation_2",
        "rotation_3",
        "weight",
        "desired_qty",
    ])
    ws.append([
        "P-100001",
        "Sample Product",
        300,
        200,
        150,
        True,
        False,
        False,
        1.250,
        12,
    ])

    info = wb.create_sheet(title="Instructions")
    info["A1"] = "Product Catalogue Upload Template"
    info["A2"] = "Units"
    info["B2"] = "Metric only"
    info["A3"] = "Dimensions"
    info["B3"] = "product_length, product_width, product_height in mm"
    info["A4"] = "Weight"
    info["B4"] = "weight in kg"
    info["A5"] = "Desired quantity"
    info["B5"] = "desired_qty is units needed per packaging analysis"
    info["A6"] = "Rotations"
    info["B6"] = "Use TRUE/FALSE, 1/0, YES/NO"
    info["A7"] = "Catalogue"
    info["B7"] = catalogue.name

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"product_catalogue_template_{catalogue.pk}.xlsx"

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def export_product_catalogue_excel(request, catalogue_id):
    catalogue = get_object_or_404(ProductCatalogue, pk=catalogue_id)
    products = catalogue.products.all().order_by("product_id")

    wb = Workbook()
    ws = wb.active
    ws.title = "Product Catalogue Export"

    ws.append([
        "product_id",
        "product_name",
        "product_length",
        "product_width",
        "product_height",
        "rotation_1",
        "rotation_2",
        "rotation_3",
        "weight",
        "desired_qty",
        "product_volume",
        "product_picture",
    ])

    for p in products:
        ws.append([
            p.product_id,
            p.product_name,
            float(p.product_length) if p.product_length is not None else None,
            float(p.product_width) if p.product_width is not None else None,
            float(p.product_height) if p.product_height is not None else None,
            p.rotation_1,
            p.rotation_2,
            p.rotation_3,
            float(p.weight) if p.weight is not None else None,
            p.desired_qty,
            float(p.product_volume) if p.product_volume is not None else None,
            p.product_picture.url if p.product_picture else "",
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"product_catalogue_export_{catalogue.pk}.xlsx"

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def upload_product_images_zip(request, catalogue_id):
    catalogue = get_object_or_404(ProductCatalogue, id=catalogue_id)

    if request.method == "POST":
        form = ProductImagesZipUploadForm(request.POST, request.FILES)
        if form.is_valid():
            zf = form.cleaned_data["file"]
            try:
                with zipfile.ZipFile(zf) as z:
                    zip_names = {name.lower(): name for name in z.namelist() if not name.endswith("/")}

                    matched = 0
                    for p in catalogue.products.all():
                        if p.product_picture:
                            continue

                        candidates = []
                        if p.product_id:
                            candidates += [
                                f"{p.product_id}.png",
                                f"{p.product_id}.jpg",
                                f"{p.product_id}.jpeg",
                                f"{p.product_id}.webp",
                            ]
                        if p.product_name:
                            safe = str(p.product_name).strip()
                            candidates += [
                                f"{safe}.png",
                                f"{safe}.jpg",
                                f"{safe}.jpeg",
                                f"{safe}.webp",
                            ]

                        found_key = None
                        for c in candidates:
                            if c.lower() in zip_names:
                                found_key = zip_names[c.lower()]
                                break

                        if not found_key:
                            continue

                        data = z.read(found_key)
                        p.product_picture.save(found_key.split("/")[-1], ContentFile(data), save=True)
                        matched += 1

                return render(
                    request,
                    "product_catalogue/upload_images_zip.html",
                    {
                        "form": ProductImagesZipUploadForm(),
                        "catalogue": catalogue,
                        "success": f"ZIP processed successfully. Images matched and saved: {matched}.",
                    },
                )
            except zipfile.BadZipFile:
                return render(
                    request,
                    "product_catalogue/upload_images_zip.html",
                    {
                        "form": form,
                        "catalogue": catalogue,
                        "error": "Invalid ZIP file.",
                    },
                )
    else:
        form = ProductImagesZipUploadForm()

    return render(
        request,
        "product_catalogue/upload_images_zip.html",
        {"catalogue": catalogue, "form": form},
    )